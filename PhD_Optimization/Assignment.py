##############################################################################       
### Import packages
from gurobipy import *
import numpy as np 
import numpy 
import pandas as pd 
from pandas import Series, ExcelWriter
import pandas
from openpyxl import load_workbook
import xlsxwriter
from numpy import linalg
from decimal import Decimal
import time
import math
import heapq
import random,string   

from Optimization import Assignment
from Simulation import Initialization
from FCFS import FCFS

##############################################################################       
### Inputs
Time = 0
TH = 180                                               ## Time horizon
THp = 5
Timefence = 60                                         ## using for fixing assigned resources (as a constraint in the optimization)     
Threshold = 0                                          ## using to specify assigned resources (post-processing) 
epsilon = 0.5                                           ## total waiting time shouldn't be more than 15 minuts       

S_be = 50                                              ## Average service time for cleaning a bed 
TNE = 4
NE= 2                                                  ## Number of EVS Staff in each shift
NT= 2                                                  ## Number of Transport Staff

NAPD = 10                                              ## Number of admitted patients per day
Chance = 0.3                                           ## Chance of admitting each patient to IU
N_days = 1
Np = int(math.ceil(float((NAPD/Chance)* N_days)))      ## Total number of patients come to ED in each day
L = (1440*N_days)/Np                                   ## Average rate of arrival time of patients to ED

G_chance = [0.5,0.5]                                   ## Gender chance [female, male]
P_priority = [1,1.1,1.2]                               ## Priority weight of patients based on their preferred IU => [1,1,1] or [1, 1.1, 1.2]
LOS = 48* 60                                           ## Average LOS
#S_b = LOS                                           
Nb= 100                                                ## Total Number of beds
NR = int(math.ceil (Nb*2/float(3)))                    ## Total Number of Rooms
IU_types = 3                                           ## IU 3 has more critical patients
NR_IU = [0.6,0.2,0.2]                                  ## chance of a room belongs to IU1 or IU2
P_R_capacity = [0.5,0.5]                               ## Probablity of capacity of room [p(capacity=1),p(capacity=2)]
Nb_IU = [0.6,0.2,0.2]
Np_IU = [0.6,0.2,0.2]
bedstates = ['occupied','dirty', 'clean']
Bstatus_chance = [0.5, 0.25, 0.25]

Bed_Utlztn = [((NAPD*Np_IU[0])/((Nb*Nb_IU[0])/(LOS/float(24*60)))), ((NAPD*Np_IU[1])/((Nb*Nb_IU[1])/(LOS/float(24*60)))), ((NAPD*Np_IU[1])/((Nb*Nb_IU[1])/(LOS/float(24*60))))]

Balancing = [1]*NE  ## coeeficient for EVS staffs workload balancing coefficients
EVS_Shift = 8*60
E_Shift_Points = [0]
for i in range (int((N_days*24*60/EVS_Shift)+1)): E_Shift_Points.append((i+1)*8*60)

IU_ED  = [np.random.normal (5,1), np.random.normal (15,2), np.random.normal (25,2)]   # distance of each IU to ED 
IU2IU = [np.random.normal (10,1), np.random.normal (20,2), np.random.normal (10,2)]    # distance of each 2 IUs: IU1_IU2, IU1_IU3, IU2-IU3

Events = ['new_patient', 'new_bed', 'new_EVS', 'new_tran']

##############################################################################       
############### Initial Data
#P, IU_Patients, Rooms, Beds, All_EVS, Tran, IU_D = Initialization(Np, N_days, L, Chance, G_chance, NR, TNE, NE, EVS_Shift, NT, IU_types, Np_IU, NR_IU, IU_ED,IU2IU, S_be, bedstates, Bstatus_chance, P_R_capacity, P_priority, LOS, Bed_Utlztn)   # P = patients in queue for assignment

##### Read Inputs from excel
Information = pd.read_excel('Information7_Reactive.xlsx', sheet_name=['IU_Patients','Beds','Rooms','All_EVS','Tran','IU_D'], na_values='n/a')
P = pd.DataFrame(index=range(0), columns=['ID','Gender','Admission time','Priority', 'IU','IU_D','Ct_p', 'LOS','Discharge time'])  # Patients in queue
IU_Patients = Information['IU_Patients']                        
Beds = Information['Beds']                  
Rooms = Information['Rooms']                 
All_EVS = Information['All_EVS']         
EVS = All_EVS[:NE]           
Tran = Information['Tran']                       
IU_D = Information['IU_D']    
#########################
    
IU_original = IU_Patients.copy(deep=True) 
IU_original = pd.DataFrame(IU_original)

Results = pd.DataFrame(index=range(len(IU_original)), columns=['Patient.ID','Admission time','Priority','Bed.ID','B.IU','d_Bed','EVS.ID','d_EVS','Transport.ID','d_Transport','tb','tc','te','tt','t_bp','T_P','LOS','Discharge time','Time_Now'])
Results.loc[:,'Patient.ID'] = IU_original.loc[:,'ID']
Results.loc[:,'Admission time'] = IU_original.loc[:,'Admission time']
Results.loc[:,'Discharge time'] = IU_original.loc[:,'Discharge time']

OPTsetting = pd.DataFrame(index=range(0), columns=['#_Run','Event_time','N_P','N_EVS','N_UB','N_Tr','N_cleaning_beds','N_Vars','N_Constrs','N_iterations','Run_time','Gap'])

##############################################################################       
###############  Start time
# funtion to write result
def multiple_dfs(df_list, sheets, file_name, spaces):
    writer = pd.ExcelWriter(file_name,engine='xlsxwriter')   
    row = 0
    for dataframe in df_list:
        dataframe.to_excel(writer,sheet_name=sheets,startrow=row , startcol=0)   
        row = row + len(dataframe.index) + spaces + 1
    writer.save()

#### Update Patients in queue for assignment 
firstP = []
for i in range(len(IU_Patients)):
    if (IU_Patients.loc[i,'Admission time'] <= Time+THp): firstP.append(i)                
    
for i in range(len(firstP)): 
    P.loc[len(P),:] = IU_Patients.loc[firstP[i],:]

IU_Patients.drop(firstP, inplace =True) 
IU_Patients = IU_Patients.reset_index()
IU_Patients = IU_Patients.drop(['index'], axis=1) 

Tp_prev = pd.DataFrame(index=range(len(P)), columns=['Patient.ID','Tp_prev'])   ## for saving previous iteration arrival time of patients
for i in range(len(P)):
    Tp_prev.loc[i,'Patient.ID'] = P.loc[i,'ID']
    
PB_prev = pd.DataFrame(index=range(0), columns=['Patient.ID','Bed.ID'])         ## previous iteration patient to bed assignment
P_FixedinTF = []
#Tp_prev = pd.DataFrame(Tp_prev)
################################################################
#writer = pd.ExcelWriter('Reports.xlsx', engine='xlsxwriter')
itr = 0   
AllInputs = dict()
AllOPTReports = dict()
AllFCFSReports = dict()

workbook = xlsxwriter.Workbook('Optimization Inputs.xlsx')
writer = ExcelWriter('Optimization Inputs.xlsx')   # Creating Excel Writer 
writer.save()

workbook = xlsxwriter.Workbook('FCFS Reports.xlsx')
writer = ExcelWriter('FCFS Reports.xlsx')   # Creating Excel Writer Object from Pandas  
writer.save()

workbook = xlsxwriter.Workbook('Optimization Reports.xlsx')
writer = ExcelWriter('Optimization Reports.xlsx')
writer.save()

EVS_allReports = pd.DataFrame()
Beds_Results = pd.DataFrame(index=range(0), columns=['Bed.ID','S.T.Occupancy','E.T.Occupancy','Patient.ID','Cleaning.S.T','Cleaning.E.T','EVS.ID'])

while ((len(IU_Patients)>0 or len(P)>0)): #and(Time <22000)
    itr = itr +1    
    if len (P) == 0: 
        Time = IU_Patients.loc[IU_Patients.loc[:,'Admission time'][0:len(IU_Patients)].idxmin(IU_Patients)]['Admission time']
        Upcoming_P = []
        for i in range(len(IU_Patients)):
            if (Time <= IU_Patients.loc[i,'Admission time'] <= Time + THp): Upcoming_P.append(i)                    
        N_next_P = len(Upcoming_P)     
        next_P = numpy.zeros(shape=(N_next_P,9))
        next_P = pd.DataFrame(next_P)
        next_P.columns = ['ID','Gender','Admission time','Priority', 'IU','IU_D','Ct_p', 'LOS', 'Discharge time']        
        for i in range(len(Upcoming_P)): 
            next_P.loc[i,:] = IU_Patients.loc[Upcoming_P[i],:]     
        P = next_P      
        P = P.reset_index()
        P = P.drop(['index'], axis=1)                               
        ### Updating upcoming IU_patients information
        IU_Patients.drop(Upcoming_P, inplace =True) 
        IU_Patients = IU_Patients.reset_index()
        IU_Patients = IU_Patients.drop(['index'], axis=1) 
        
        ### Update Beds-Gender  based on the Time      
        for b in range(len(Beds)):
             if Beds.loc[b,'Availability time'] <= Time:
                 Beds.loc[b,'P_Gender'] = 0
                 
        ### Update Rooms-Gender
        for r in range(len(Rooms)):
            rb = np.where(Beds['Room'] == r)[0].tolist()
            rbidx = 0
            for b in rb:
                if Beds.loc[b,'P_Gender']!= Rooms.loc[r,'Gender']:
                    rbidx = rbidx+1
            if rbidx == Rooms.loc[r,'Capacity']:  
                if len(np.where(Beds.iloc[rb]['P_Gender'] != 0)[0].tolist()) >0 :
                    gbidx = np.where(Beds.iloc[rb]['P_Gender'] != 0)[0].tolist()
                    Rooms.loc[r,'Gender'] = Beds.loc[rb[gbidx[0]],'P_Gender']
                else: Rooms.loc[r,'Gender'] = 0
    ###################
    ## EVS Staff selection based on shifts
    EVS = pd.concat([EVS, All_EVS[(All_EVS['END_Shift'] > Time) & (All_EVS['Availability time'] <= Time + TH )]]) #            All_EVS.drop(All_EVS.index[:len(EVS)], inplace =True) 
    EVS = EVS.drop_duplicates(subset=['EVS.ID', 'END_Shift'], keep='first')

    EVS = EVS[EVS.END_Shift > Time]
    EVS = EVS.reset_index()
    EVS = EVS.drop(['index'], axis=1)    
    ##############################################################################       
    ############### Optimization Results
    ##### Write Inputs to excel
    writer = pd.ExcelWriter('optInputs.xlsx')    
    P.to_excel(writer,'P')                         # Save patients information as input
    Beds.to_excel(writer,'Beds')                   # Save Beds information as input
    Rooms.to_excel(writer,'Rooms')                 # Save Rooms information as input
    EVS.to_excel(writer,'EVS')                     # Save EVS information as input
    Tran.to_excel(writer,'Tran')                   # Save Tran information as input
    IU_original.to_excel(writer,'IU_original')     # Save IU_original patients information as input   
    IU_D.to_excel(writer,'IU_D')                   # Save IUs information as input
    writer.save()

    ##### Read Inputs from excel
    AllInputs[itr] = pd.read_excel('optInputs.xlsx', sheet_name=['P','Beds','Rooms','EVS','Tran','IU_original','IU_D'], na_values='n/a')
    P = AllInputs[itr]['P']                        
    Beds = AllInputs[itr]['Beds']                  
    Rooms = AllInputs[itr]['Rooms']                 
    EVS = AllInputs[itr]['EVS']                    
    Tran = AllInputs[itr]['Tran']                   
    IU_original = AllInputs[itr]['IU_original']     
    IU_D = AllInputs[itr]['IU_D']      
    
    ###################################################
    ## Write all Inputs in excel
    sheet_name0 = 'OPT-Input %d'%itr        
  
    book = load_workbook('Optimization Inputs.xlsx')
    writer = pandas.ExcelWriter('Optimization Inputs.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    P.to_excel(writer, sheet_name0, index=False, startrow=0 , startcol=0)
    Beds.to_excel(writer, sheet_name0, index=False, startrow=len(P)+2 , startcol=0)
    Rooms.to_excel(writer, sheet_name0, index=False, startrow=len(P)+len(Beds)+4 , startcol=0)
    EVS.to_excel(writer, sheet_name0, index=False, startrow=len(P)+len(Beds)+len(Rooms)+6 , startcol=0)
    Tran.to_excel(writer, sheet_name0, index=False, startrow=len(P)+len(Beds)+len(Rooms)+len(EVS)+8 , startcol=0)
  
    writer.save()                        
    ###########################      
    Assignments_ID = []
    Nbc = 0 
    RunTimelimit = 600
    OptGap = []
    AllBeta = []
    Allobjval = []
    Run_T = []  
    try:
        itry = 0
        Numberoftry4OPT = 0
        objval1_prev = 0
        objval = [0,0,0,0,0]
        Assignments_0 = pd.DataFrame()
        Assignments_ID, EVS_Report, Tran_Report, Bed_Report, Bed_EVS_assign, D, useful_Beds, objval , opt_Gap, H, h_ip, N_Vars, N_Constrs, runtime, beta, Numberoftry4OPT = Assignment(Numberoftry4OPT, itry, objval1_prev, objval, epsilon, Assignments_0, Tp_prev, PB_prev, P_FixedinTF, P, Beds, Rooms, EVS, NE, Tran, Time, IU_original, IU_D, IU_types, LOS, Balancing, Nbc, RunTimelimit)    
        OptGap.append(opt_Gap)
        Run_T.append(runtime)
        AllBeta.append([round(x,5) for x in beta[0]])
        Allobjval.append([round(x,5) for x in objval])
        Assignments_0 = Assignments_ID.copy(deep=True)   #### for initial solution in optimization model
        objval1_new = objval[0]
        objval1_prev = objval[0]
        Nbc0 = len(Bed_EVS_assign)
        if len(EVS)== NE:   Nbc_final = max(Nbc0, 3*NE)
        else:               Nbc_final = max(Nbc0, 3*NE+1) #3*NE+2
        Nbc = len(Bed_EVS_assign)+ 1
        while (Nbc <= Nbc_final):
            itry = itry + 1     
            Assignments_ID, EVS_Report, Tran_Report, Bed_Report, Bed_EVS_assign, D, useful_Beds, objval , opt_Gap, H, h_ip, N_Vars, N_Constrs, runtime, beta, Numberoftry4OPT = Assignment(Numberoftry4OPT, itry, objval1_prev, objval, epsilon, Assignments_0, Tp_prev, PB_prev, P_FixedinTF, P, Beds, Rooms, EVS, NE, Tran, Time, IU_original, IU_D, IU_types, LOS, Balancing, Nbc, RunTimelimit)    
            OptGap.append(opt_Gap)
            Run_T.append(runtime)
            AllBeta.append([round(x,5) for x in beta[1]])
            Allobjval.append([round(x,5) for x in objval])
            Assignments_0 = Assignments_ID.copy(deep=True)
            objval1_new = objval[0]
            Nbc = len(Bed_EVS_assign)+ 1    
            RunTimelimit = 300       
    except:
        pass
    
    ##############################################################################       
    ############### Optimization Settings 
    index_OPTsetting = len(OPTsetting)
    OPTsetting.loc[index_OPTsetting, '#_Run'] = itr
    OPTsetting.loc[index_OPTsetting, 'Event_time'] = Time                   
    OPTsetting.loc[index_OPTsetting, 'N_P'] = len(P)
    OPTsetting.loc[index_OPTsetting, 'N_EVS'] = len(EVS)                   
    OPTsetting.loc[index_OPTsetting, 'N_UB'] = len(useful_Beds)   
    OPTsetting.loc[index_OPTsetting, 'N_Tr'] = len(Tran) 
    OPTsetting.loc[index_OPTsetting, 'N_cleaning_beds'] = len(Bed_EVS_assign)               
    OPTsetting.loc[index_OPTsetting, 'N_iterations'] = itry  
    OPTsetting.loc[index_OPTsetting, 'Run_time'] = Run_T               
    OPTsetting.loc[index_OPTsetting, 'Gap'] = OptGap       
    OPTsetting.loc[index_OPTsetting, 'N_Vars'] = N_Vars                
    OPTsetting.loc[index_OPTsetting, 'N_Constrs'] = N_Constrs                        
        
    ###########################
    for i in range(len(Assignments_ID)):  
        Results.loc[int(np.where(Results['Patient.ID'] == str(Assignments_ID.loc[i,'Patient.ID']))[0]),0:(Results.shape[1]-1)] = Assignments_ID.loc[i,:]
        Results.loc[int(np.where(Results['Patient.ID'] == Assignments_ID.loc[i,'Patient.ID'])[0]),'Time_Now'] = Time
        Results.loc[int(np.where(Results['Patient.ID'] == Assignments_ID.loc[i,'Patient.ID'])[0]),'Ndb'] = len(Beds.loc[(Beds['Availability time'] <= Time) & (Beds['Status']=='dirty')]) ## number of available dirty beds
        Results.loc[int(np.where(Results['Patient.ID'] == Assignments_ID.loc[i,'Patient.ID'])[0]),'Waiting_T'] = Assignments_ID.loc[i,'T_P']-Assignments_ID.loc[i,'Admission time']- Assignments_ID.loc[i,'S_tp']
        Results.loc[int(np.where(Results['Patient.ID'] == Assignments_ID.loc[i,'Patient.ID'])[0]),'B.IU'] = Assignments_ID.loc[i,'B.IU']
        Results.loc[int(np.where(Results['Patient.ID'] == Assignments_ID.loc[i,'Patient.ID'])[0]),'LOS'] = Assignments_ID.loc[i,'LOS']
        Results.loc[int(np.where(Results['Patient.ID'] == Assignments_ID.loc[i,'Patient.ID'])[0]),'Priority'] = Assignments_ID.loc[i,'Priority']

    # Save list of Reports
    df1 = pd.DataFrame([{'Time_now': Time,'Run_Time': Run_T, 'opt_Gap':OptGap, 'opt_objVal':objval,'N_Constrs':N_Constrs,'N_Vars':N_Vars,'beta':AllBeta,'All_objval':Allobjval, 'Numberoftry4OPT':Numberoftry4OPT}])
    df2 = pd.DataFrame([{'min.Waiting.T':min(Assignments_ID['Waiting.T'])  ,'max.Waiting.T':max(Assignments_ID['Waiting.T'])  ,'Average.Waiting.T':sum(Assignments_ID['Waiting.T'])/len(Assignments_ID['Waiting.T'])}])
    resultInfo = pd.concat([df1, df2], axis=1, join='inner')
    Duplications = pd.DataFrame([{'Duplications':D}])
    h_ip = pd.DataFrame(h_ip)
    Reports = [resultInfo, Duplications, Assignments_ID, EVS_Report, Tran_Report, Bed_Report, IU_D, h_ip, H, useful_Beds]
    
    multiple_dfs(Reports, 'Report', 'Report.xlsx', 1)
        

    ####### Save the results
    AllOPTReports[itr] = Reports
    
    sheet_name = 'OPT-Report %d'%itr        
  
    book = load_workbook('Optimization Reports.xlsx')
    writer = pandas.ExcelWriter('Optimization Reports.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    resultInfo.to_excel(writer, sheet_name, index=False, startrow=0 , startcol=0)
    Duplications.to_excel(writer, sheet_name, index=False, startrow=len(resultInfo)+2 , startcol=0)
    Assignments_ID.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+4 , startcol=0)
    EVS_Report.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+len(Assignments_ID)+6 , startcol=0)
    Tran_Report.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+len(Assignments_ID)+len(EVS_Report)+8 , startcol=0)
    Bed_Report.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+len(Assignments_ID)+len(EVS_Report)+len(Tran_Report)+10 , startcol=0)
    IU_D.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+len(Assignments_ID)+len(EVS_Report)+len(Tran_Report)+len(Bed_Report)+12 , startcol=0)
    h_ip.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+len(Assignments_ID)+len(EVS_Report)+len(Tran_Report)+len(Bed_Report)+len(IU_D)+14 , startcol=0)
    H.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+len(Assignments_ID)+len(EVS_Report)+len(Tran_Report)+len(Bed_Report)+len(IU_D)+len(h_ip)+16 , startcol=0)
    useful_Beds.to_excel(writer, sheet_name, index=False, startrow=len(Duplications)+len(resultInfo)+len(Assignments_ID)+len(EVS_Report)+len(Tran_Report)+len(Bed_Report)+len(IU_D)+len(h_ip)+len(H)+18 , startcol=0)
    
    writer.save()
     
     ### Save EVS Reports
    EVS_Report['itr'] = itr
    if itr > 1:
        EVS_Report = EVS_Report.reset_index(drop=True)
        EVS_previous = EVS_allReports.copy(deep=True)         
        EVS_previous = EVS_previous.reset_index(drop=True)
        for ir in range(len(EVS_Report)):
            bedname = str(EVS_Report['Bed.ID'][ir])
            EVSindlist1 = np.where(np.logical_and(EVS_previous['Bed.ID'] == bedname, EVS_previous['Bed.EA.Time'] == EVS_Report['Bed.EA.Time'][ir]))[0]   ## remove the rows that have the same beds in each time ( prevent cleaning one bed for more than 1 time)
            EVS_previous.drop(EVSindlist1, inplace =True)           
            
            EVS_previous = EVS_previous.reset_index(drop=True)
           
        EVS_allReports = EVS_previous.copy(deep=True)       
    EVS_allReports = EVS_allReports.append(EVS_Report)
    EVS_allReports = EVS_allReports.reset_index(drop=True)

    for ne in range(NE):
        Balancing[ne] = len(EVS_allReports[EVS_allReports['EVS.ID'] == EVS.loc[ne,'EVS.ID']])
           
    ##############################################################################       
    ############### FCFS Results
    FCFS_Beds, FCFS_P, FCFS_objVal1, FCFS_objVal2 = FCFS(P, Beds, Rooms, EVS, Tran, Time, IU_original, IU_D, useful_Beds)   
    FCFSInfo = pd.DataFrame([{'FCFS_objVal1': FCFS_objVal1 , 'FCFS_objVal2': FCFS_objVal2, 'Min.W.T':min(FCFS_P['Waiting.T']), 'Max.W.T':max(FCFS_P['Waiting.T']),'Mean.W.T':sum(FCFS_P['Waiting.T'])/len(FCFS_P['Waiting.T'])}])
    
    FCFS_Reports =[FCFSInfo, FCFS_Beds, FCFS_P]
    multiple_dfs(FCFS_Reports, 'FCFS_Report', 'FCFS_Report.xlsx', 2)
    
    ####### Save the results
    AllFCFSReports[itr] = FCFS_Reports
    
    sheet_name1 = 'FCFS-Report %d'%itr        
  
    book = load_workbook('FCFS Reports.xlsx')
    writer = pandas.ExcelWriter('FCFS Reports.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    FCFSInfo.to_excel(writer, sheet_name1, index=False, startrow=0 , startcol=0)
    FCFS_Beds.to_excel(writer, sheet_name1, index=False, startrow=len(FCFSInfo)+2 , startcol=0)
    FCFS_P.to_excel(writer, sheet_name1, index=False, startrow=len(FCFS_Beds)+len(FCFSInfo)+4 , startcol=0)
    
    writer.save()
          
    #############################################
    ############### Next Event Time
    if len(IU_Patients)>0 : min_IUP = IU_Patients.loc[IU_Patients.loc[:,'Admission time'][0:len(IU_Patients)].idxmin(IU_Patients)]['Admission time']

    if len(P)>0 : min_Pq = P.loc[P.loc[:,'Admission time'][0:len(P)].idxmin(P)]['Admission time']
   
    if len(Beds.loc[(Beds['Availability time'] > Time)])>0: min_Bed = min(Beds.loc[(Beds['Availability time'] > Time)]['Availability time'])    
    else:    min_Bed = Time
    
    min_EVS1 = min(EVS.loc[(EVS['Availability time'] >= 0)]['Availability time']) 
    if len(EVS_Report)>0 : min_EVS2 = min(EVS_Report.loc[(EVS_Report['Cleaning.E.T'] >= 0)]['Cleaning.E.T']) 
    
    min_Tran = min(Tran.loc[(Tran['Availability time'] >= 0)]['Availability time']) 
    
    if len(EVS_Report)>0 : min_Times = [min_IUP, min_Pq, min_Bed, min_EVS1, min_EVS2, min_Tran] 
    else :    min_Times = [min_IUP, min_Pq, min_Bed, min_EVS1, min_Tran]             
                        
    TimeNow = Time
    Time = max(TimeNow+30, min([x for x in min_Times if (x > TimeNow)]))    ## Time of next event, atleast 30 minutes difference is between two time events
    
    ##############################################################################  
    ### information needed for next iteration    
    P_FixedinTF = []
    Tp_prev = pd.DataFrame(index=range(len(Assignments_ID)), columns=['Patient.ID','Tp_prev'])
    PB_prev = pd.DataFrame(index=range(len(Assignments_ID)), columns=['Patient.ID','Bed.ID'])  ## previous iteration patient to bed assignment
    for i in range(len(Assignments_ID)):  
        Tp_prev.loc[i,'Patient.ID'] = Assignments_ID.loc[i,'Patient.ID']
        Tp_prev.loc[i,'Tp_prev'] = Assignments_ID.loc[i,'T_P']
        PB_prev.loc[i,'Patient.ID'] = Assignments_ID.loc[i,'Patient.ID']
        PB_prev.loc[i,'Bed.ID'] = Assignments_ID.loc[i,'Bed.ID']
        if (Assignments_ID.loc[i,'T_P'] >= Time) and (Assignments_ID.loc[i,'T_P'] <= Time + Timefence ) :  P_FixedinTF.append(Assignments_ID.loc[i,'Patient.ID'])
            
    ####################################################################################################
    ############### Updates based on assignment results
    
    delta_T = Time 
    next_q = []         ## Remaining patients in queue from previous iteration 
    Assigned_P = []     ## Assigned patients that left the queue from previous iteration
    Upcoming_P = []     ## New patient from ED (IU_Patients)
    
    for i in range(len(P)):
        if  (Assignments_ID.loc[i,'T_P'] - Assignments_ID.loc[i,'S_tp']) > Time + Threshold : next_q.append(i)           ## S-D  ## 
        else: Assigned_P.append(i) 
                    
    for i in range(len(IU_Patients)):
        if (TimeNow <= IU_Patients.loc[i,'Admission time'] <= Time + THp): Upcoming_P.append(i)                

    N_next_P = len(Upcoming_P) + len(next_q)    
    next_P = numpy.zeros(shape=(N_next_P,9))
    next_P = pd.DataFrame(next_P)
    next_P.columns = ['ID','Gender','Admission time','Priority', 'IU','IU_D','Ct_p', 'LOS','Discharge time']        
            
    for i in range(len(next_q)):    
        next_P.loc[i,:] = P.loc[next_q[i],:]

    for i in range(len(Upcoming_P)): 
        next_P.loc[i+len(next_q),:] = IU_Patients.loc[Upcoming_P[i],:]
 
    P = next_P      
    P = P.reset_index()
    P = P.drop(['index'], axis=1)         
                  
    ############### Updating upcoming IU_patients information
    IU_Patients.drop(Upcoming_P, inplace =True) 
    IU_Patients = IU_Patients.reset_index()
    IU_Patients = IU_Patients.drop(['index'], axis=1) 
                                          
    ############### Updating Resources
    Assigned_Bed = []   ## the IDs of beds that are assigned in previous step and availability time should be updated
    Assigned_EVS = []
    Assigned_Tran = []

    ### Update assigned Beds to patients    
    for i in range(len(Assigned_P)): 
        Assigned_Bed.append(str(Assignments_ID.loc[Assigned_P[i],'Bed.ID']))
        Beds.loc[int(np.where(Beds['Bed.ID'] == Assignments_ID.loc[Assigned_P[i],'Bed.ID'])[0]),'Availability time'] = Assignments_ID.loc[Assigned_P[i],'Discharge time'] #Assignments_ID.loc[Assigned_P[i],'LOS'] + Assignments_ID.loc[Assigned_P[i],'T_P']
        Beds.loc[int(np.where(Beds['Bed.ID'] == Assignments_ID.loc[Assigned_P[i],'Bed.ID'])[0]),'Status'] = 'occupied'
        Beds.loc[int(np.where(Beds['Bed.ID'] == Assignments_ID.loc[Assigned_P[i],'Bed.ID'])[0]),'P_Gender'] = IU_original.loc[int(np.where(IU_original['ID'] == Assignments_ID.loc[Assigned_P[i],'Patient.ID'])[0]),'Gender']

        bii = len(Beds_Results)
        Beds_Results.loc[bii, 'Bed.ID'] = Assignments_ID.loc[Assigned_P[i],'Bed.ID']
        Beds_Results.loc[bii, 'S.T.Occupancy'] = Assignments_ID.loc[Assigned_P[i],'T_P']
        Beds_Results.loc[bii, 'E.T.Occupancy'] = Assignments_ID.loc[Assigned_P[i],'Discharge time']#Assignments_ID.loc[Assigned_P[i],'LOS'] + Assignments_ID.loc[Assigned_P[i],'T_P']
        Beds_Results.loc[bii, 'Patient.ID'] = Assignments_ID.loc[Assigned_P[i],'Patient.ID']
        
    EVS_Report = EVS_Report.reset_index()
    EVS_Report = EVS_Report.drop(['index'], axis=1)    
    for i in range(len(EVS_Report)):
        if (not any(EVS_Report.loc[i,'Bed.ID'] in s for s in Assigned_Bed)):
            if EVS_Report.loc[i,'Traveling.S.T'] <= Time :
                Beds.loc[int(np.where(Beds['Bed.ID'] == EVS_Report.loc[i,'Bed.ID'])[0]),'Status'] = 'clean'
                Beds.loc[int(np.where(Beds['Bed.ID'] == EVS_Report.loc[i,'Bed.ID'])[0]),'P_Gender'] = 0
                Beds.loc[int(np.where(Beds['Bed.ID'] == EVS_Report.loc[i,'Bed.ID'])[0]),'Availability time'] = EVS_Report.loc[i,'Cleaning.E.T']

                eii = len(Beds_Results)
                Beds_Results.loc[eii, 'Bed.ID'] = EVS_Report.loc[i,'Bed.ID']
                Beds_Results.loc[eii, 'Cleaning.S.T'] = EVS_Report.loc[i,'Cleaning.S.T']
                Beds_Results.loc[eii, 'Cleaning.E.T'] = EVS_Report.loc[i,'Cleaning.E.T']
                Beds_Results.loc[eii, 'EVS.ID'] = EVS_Report.loc[i,'EVS.ID']

    for b in range(len(Beds)):
         if Beds.loc[b,'Availability time'] <= Time:
             Beds.loc[b,'P_Gender'] = 0
             if Beds.loc[b,'Status'] == 'occupied' : Beds.loc[b,'Status'] = 'dirty'
                          
    ### Update Rooms-Gender
    for r in range(len(Rooms)):
        rb = np.where(Beds['Room'] == r)[0].tolist()
        rbidx = 0
        for b in rb:
            if Beds.loc[b,'P_Gender']!= Rooms.loc[r,'Gender']:
                rbidx = rbidx+1
        if rbidx == Rooms.loc[r,'Capacity']:  
            if len(np.where(Beds.iloc[rb]['P_Gender'] != 0)[0].tolist()) > 0 :
                gbidx = np.where(Beds.iloc[rb]['P_Gender'] != 0)[0].tolist()
                Rooms.loc[r,'Gender'] = Beds.loc[rb[gbidx[0]],'P_Gender']
            else: Rooms.loc[r,'Gender'] = 0
            
    ### Update Transport staff        
    for i in range(len(Assigned_P)): 
        Assigned_Tran.append(Assignments_ID.loc[Assigned_P[i],'Transport.ID'])
        Tran_prevTime = Tran.loc[int(np.where(Tran['Tran.ID'] == Assignments_ID.loc[Assigned_P[i],'Transport.ID'])[0]),'Availability time']
        Tran.loc[int(np.where(Tran['Tran.ID'] == Assignments_ID.loc[Assigned_P[i],'Transport.ID'])[0]),'Availability time'] = max(Tran_prevTime, Assignments_ID.loc[Assigned_P[i],'T_P'] + IU_original.loc[int(np.where(IU_original['ID'] == Assignments_ID.loc[Assigned_P[i],'Patient.ID'])[0]),'IU_D'])

    ### Update EVS   
    for k in range(len(EVS_Report)):
        if EVS_Report.loc[k,'Traveling.S.T'] <= Time :
            EVS.loc[int(np.where(EVS['EVS.ID'] == EVS_Report.loc[k,'EVS.ID'])[0]),'Availability time'] = EVS_Report.loc[k,'Cleaning.E.T']
            EVS.loc[int(np.where(EVS['EVS.ID'] == EVS_Report.loc[k,'EVS.ID'])[0]),'Location'] = EVS_Report.loc[k,'B.IU']
            
    ###############################################################################################
    ### Transport Results            
    Trans_Results = pd.DataFrame(index=range(len(IU_original)), columns=['Patient.ID','Admission time','IU','IU_D','Transport.ID','d_Transport','availability_time','start_time','T_P', 'Waiting_time'])
    Trans_Results.loc[:,['Patient.ID', 'Admission time']] = Results.loc[:,['Patient.ID', 'Admission time']]
    Trans_Results.loc[:,['IU','IU_D']] = IU_original.loc[:,['IU','IU_D']]
    Trans_Results.loc[:,['Transport.ID','d_Transport']] = Results.loc[:,['Transport.ID','d_Transport']]
    Trans_Results.loc[:,'availability_time'] = Results.loc[:,'tt']
    Trans_Results.loc[:,'T_P'] = Results.loc[:,'T_P']
    Trans_Results.loc[:,'start_time'] = Trans_Results.loc[:,'T_P'] - Trans_Results.loc[:,'IU_D']
    Trans_Results.loc[:,'Waiting_time'] = Trans_Results.loc[:,'start_time'] - Trans_Results.loc[:,'availability_time']
    Trans_Results = Trans_Results.sort_values('T_P',ascending=True)
    
    ### EVS Results            
    EVS_Results1 = EVS_allReports.drop_duplicates(subset=['EVS.ID', 'Bed.ID', 'Cleaning.E.T'], keep='last')
    EVS_Results1 = EVS_Results1.sort_values('itr',ascending=True)
    EVS_Results1 = EVS_Results1.reset_index(drop=True)
    EVS_Results = EVS_Results1.drop_duplicates(subset=['EVS.ID', 'Earlieast.Availability.T'], keep='last')
    EVS_Results = EVS_Results.sort_values('EVS.ID',ascending=True)
    EVS_Results = EVS_Results.reset_index(drop=True)
    
    ##### Write to excel summary of results
    writer = pd.ExcelWriter('Optimization Results.xlsx') 
    IU_original.to_excel(writer,'IU_original') 
    Results.to_excel(writer,'Patients Assignments')  
    Trans_Results.to_excel(writer,'Trans_Results')    
    EVS_Results.to_excel(writer,'EVS_Results') 
    Beds_Results.to_excel(writer,'Beds_Results')       
    Beds.to_excel(writer,'Beds_Last status')
    OPTsetting.to_excel(writer,'OPTsetting')
    writer.save()
    
writer.save()
###############################################################################################
### Transport Results            
Trans_Results = pd.DataFrame(index=range(len(IU_original)), columns=['Patient.ID','Admission time','IU','IU_D','Transport.ID','d_Transport','availability_time','start_time','T_P', 'Waiting_time'])
Trans_Results.loc[:,['Patient.ID', 'Admission time']] = Results.loc[:,['Patient.ID', 'Admission time']]
Trans_Results.loc[:,['IU','IU_D']] = IU_original.loc[:,['IU','IU_D']]
Trans_Results.loc[:,['Transport.ID','d_Transport']] = Results.loc[:,['Transport.ID','d_Transport']]
Trans_Results.loc[:,'availability_time'] = Results.loc[:,'tt']
Trans_Results.loc[:,'T_P'] = Results.loc[:,'T_P']
Trans_Results.loc[:,'start_time'] = Trans_Results.loc[:,'T_P'] - Trans_Results.loc[:,'IU_D']
Trans_Results.loc[:,'Waiting_time'] = Trans_Results.loc[:,'start_time'] - Trans_Results.loc[:,'availability_time']
Trans_Results = Trans_Results.sort_values('T_P',ascending=True)

### EVS Results            
EVS_Results1 = EVS_allReports.drop_duplicates(subset=['EVS.ID', 'Bed.ID', 'Cleaning.E.T'], keep='last')
EVS_Results1 = EVS_Results1.sort_values('itr',ascending=True)
EVS_Results1 = EVS_Results1.reset_index(drop=True)
EVS_Results = EVS_Results1.drop_duplicates(subset=['EVS.ID', 'Earlieast.Availability.T'], keep='last')
EVS_Results = EVS_Results.sort_values('EVS.ID',ascending=True)
EVS_Results = EVS_Results.reset_index(drop=True)

##### Write to excel summary of results
writer = pd.ExcelWriter('Optimization Results.xlsx') 
IU_original.to_excel(writer,'IU_original') 
Results.to_excel(writer,'Patients Assignments')  
Trans_Results.to_excel(writer,'Trans_Results')    
EVS_Results.to_excel(writer,'EVS_Results') 
Beds_Results.to_excel(writer,'Beds_Results')       
Beds.to_excel(writer,'Beds_Last status')
OPTsetting.to_excel(writer,'OPTsetting')
writer.save()


